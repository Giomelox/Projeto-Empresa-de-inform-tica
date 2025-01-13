import time
import pandas as pd
import tkinter as tk
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
import xml.etree.ElementTree as ET
from kivy.uix.button import Button
from tkinter import filedialog
from openpyxl import Workbook
from datetime import datetime
from xml.dom import minidom
import shutil
import imaplib
import pickle
import email
import os
import re
import requests
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout


if os.name == 'nt':  # Para Windows
    download_dir = Path(os.getenv('USERPROFILE')) / 'Downloads'
else:  # Para macOS e Linux
    download_dir = Path.home() / 'Downloads'

current_dir = Path(__file__).parent

aux_path_json = current_dir / 'credentials.json'

aux_path_XML_destino = download_dir / 'XML_DELL'

aux_path_XML_destino.mkdir(parents = True, exist_ok = True)

def log_message(log_input, message):
    log_input.text += message + '\n'

class Log:
    def __init__(self):
        self.text = ''

log_input = Log() 

class SolicitarPlanilha:
    def __init__(self, log_input):
        self.log_input = log_input

    def escolher_planilha(self):
        # Iniciar uma janela oculta do tkinter
        root = tk.Tk()
        root.withdraw()  # Ocultar a janela principal

        # Abrir o diálogo de seleção de arquivo e permitir que o usuário selecione uma planilha
        file_path = filedialog.askopenfilename(
            title="Selecione a planilha",
            filetypes=[("Planilhas Excel", "*.xlsx *.xls")]
        )

        # Ler as folhas do arquivo Excel
        try:
            excel_file = pd.ExcelFile(file_path)
            sheets = excel_file.sheet_names

            # Verificar se a folha 'XML' existe
            if 'Processos devolução (programa)' in sheets:
                planilha_df = pd.read_excel(file_path, sheet_name = 'Processos devolução (programa)')

            else:
                planilha_df = pd.read_excel(file_path, sheet_name = sheets[0], header = None)

            log_message(self.log_input, f'Planilha selecionada: {file_path}\n')

            return planilha_df

        except Exception as e:
            log_message(self.log_input, f'Erro ao ler o arquivo: {str(e)}\n')
            return None

solicitar = SolicitarPlanilha(log_input)

planilha_df = solicitar.escolher_planilha()

@staticmethod
def obter_configs():
    """Carrega as configurações do arquivo e retorna como dicionário"""
    try:
        with open('email.txt', 'r') as f:
            dados = f.read()
            dados = [item.strip() for item in dados.split(',')]
            return {
                "email": dados[0],
                "usuario_Elogistic": dados[1],
                "senha_Elogistic": dados[2],
                'usuario_IOB': dados[3],
                'senha_IOB': dados[4],
                'aliquota_interna': dados[5],
                'nome_credenciada': dados[6],
                'cnpj_credenciada': dados[7],
                'caixa_emails': dados[8],
            }
    except FileNotFoundError:
        return {}

if obter_configs:
    conta_email = obter_configs().get('email')
    caixa_emails = obter_configs().get('caixa_emails')
    usuario_elogistica = obter_configs().get('usuario_Elogistic')
    senha_elogistica = obter_configs().get('senha_Elogistic')
    usuario_IOB = obter_configs().get('usuario_IOB')
    senha_IOB = obter_configs().get('senha_IOB')
    aliquota_interna = obter_configs().get('aliquota_interna')
    nome_credenciada = obter_configs().get('nome_credenciada')
    cnpj_credenciada = obter_configs().get('cnpj_credenciada')

# Baixar arquivos Dell e Valores devolução Dell
def conectar_email_e_baixar_arquivos_Dell(self, *args):
    substrings_Dell = []

    for index, row in planilha_df.iterrows():

        cell_Dell = row[1]

        if isinstance(cell_Dell, str) and len(cell_Dell) >= 34:
            # Extrai a substring e adiciona à lista
            substring_Dell = cell_Dell[28:34]
            substrings_Dell.append(substring_Dell)
        elif isinstance(cell_Dell, float):
            cell_int = int(cell_Dell)
            substrings_Dell.append(cell_int)
        else:
            substrings_Dell.append(cell_Dell)

    SCOPES_devolução = ['https://mail.google.com/']

    def get_gmail_service():
        """Realiza a autenticação OAuth2 e retorna o serviço do Gmail."""
        
        creds = None
        # Carrega as credenciais do arquivo 'token.pickle' se ele existir
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)

        # Verifica se as credenciais são válidas, e se não, tenta atualizar ou obter novas
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())  # Tenta atualizar o token
                except Exception as e:
                    log_message(self.log_input, f'Erro ao atualizar as credenciais: {e}')
                    # Se falhar, precisa obter novas credenciais
                    flow = InstalledAppFlow.from_client_secrets_file(str(aux_path_json), SCOPES_devolução)
                    creds = flow.run_local_server(port=0)  # Inicia o fluxo de autenticação
            else:
                # Se não houver credenciais, inicia o fluxo de autenticação
                flow = InstalledAppFlow.from_client_secrets_file(str(aux_path_json), SCOPES_devolução)
                creds = flow.run_local_server(port=0)

            # Salva as credenciais para futuras execuções
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)

        return creds

    def generate_oauth2_string(username, access_token):
        """Gera a string de autenticação OAuth2 no formato necessário para IMAP."""
        auth_string = f'user={username}\1auth=Bearer {access_token}\1\1'
        return auth_string.encode('ascii')

    def connect_to_gmail_imap(creds, conta_email):
        """Conecta ao Gmail via IMAP usando OAuth2."""
        imap_host = 'imap.gmail.com'
        mail = imaplib.IMAP4_SSL(imap_host)
        
        # Gera a string de autenticação OAuth2
        auth_string = generate_oauth2_string(conta_email, creds.token)
        
        # Autentica no servidor IMAP usando OAuth2
        try:
            mail.authenticate('XOAUTH2', lambda x: auth_string)
            log_message(self.log_input, '\nAutenticação bem sucedida, seguindo para baixar emails\n')
        except Exception as e:
            log_message(self.log_input, f'\nOcorreu um erro durante a autenticação com servidor: {e}')
            return None
        
        return mail

    # Inicie a autenticação
    creds_devolução = get_gmail_service()

    # Conecte ao Gmail via IMAP usando as credenciais OAuth2
    mail = connect_to_gmail_imap(creds_devolução, conta_email)

    if mail:
        # Selecionar a caixa de entrada ou outro rótulo
        mail.select(f'"{caixa_emails}"')

        try:
            processed_files_devolução = []

            idx = 1

            for xml in substrings_Dell:
                status, email_ids = mail.search(None, f'(SUBJECT "{xml}")')
                log_message(self.log_input, f'{idx} - Buscando emails com assunto: "{xml}"\n')

                for email_id in email_ids[0].split():
                    status, email_data = mail.fetch(email_id, '(RFC822)')
                    raw_email = email_data[0][1]
                    message = email.message_from_bytes(raw_email)

                    # Processar o e-mail e anexos
                    for part in message.walk():
                        if part.get_content_maintype() == 'multipart':
                            continue
                        if part.get('Content-Disposition') is None:
                            continue

                        file_name = part.get_filename()

                        # Verificar se o anexo já foi processado
                        if file_name and file_name.endswith('-procNFe.xml') and file_name not in processed_files_devolução:
                            log_message(self.log_input, f'{idx} - Anexo baixado: {file_name}\n')

                            try:
                                save_path = os.path.join(aux_path_XML_destino, file_name)
                                os.makedirs(aux_path_XML_destino, exist_ok = True)

                                # Salvar o arquivo
                                with open(save_path, 'wb') as f:
                                    f.write(part.get_payload(decode = True))

                                if os.path.exists(save_path):
                                    destino_path = os.path.join(aux_path_XML_destino, file_name)
                                    shutil.move(save_path, destino_path)

                                    # Adicionar o arquivo processado na lista
                                    processed_files_devolução.append(file_name)
                                else:
                                    log_message(self.log_input, f'{idx} - O arquivo de origem não foi encontrado: {save_path}')

                                    idx += 1

                            except Exception as e:
                                log_message(self.log_input, f'{idx} - Ocorreu um erro ao processar o anexo "{file_name}": {e}')

                                idx += 1
                idx += 1

        except Exception as e:
            log_message(self.log_input, f'Não foi possível prosseguir com a busca no email: {e}')
    else:
        log_message(self.log_input, f'A conexão IMAP falhou.')

    substrings_Dell.clear()

# Baixar arquivos HP e valores devolução HP
def conectar_email_e_baixar_arquivos_HP(self, *args):
    substrings_hp = []
    for index, row in planilha_df.iterrows():
        cell_hp = row[1]

        if isinstance(cell_hp, str) and len(cell_hp) >= 34:
            # Extrai a substring e adiciona à lista
            substring_hp = cell_hp[28:34]
            substrings_hp.append(substring_hp)
        elif isinstance(cell_hp, float):
            cell_int = int(cell_hp)
            substrings_hp.append(cell_int)
        else:
            substrings_hp.append(cell_hp)

    SCOPES = ['https://mail.google.com/']

    def get_gmail_service():
        """Realiza a autenticação OAuth2 e retorna o serviço do Gmail."""
        
        creds = None
        # Carrega as credenciais do arquivo 'token.pickle' se ele existir
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)

        # Verifica se as credenciais são válidas, e se não, tenta atualizar ou obter novas
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())  # Tenta atualizar o token
                except Exception as e:
                    log_message(self.log_input, f'Erro ao atualizar as credenciais: {e}')
                    # Se falhar, precisa obter novas credenciais
                    flow = InstalledAppFlow.from_client_secrets_file(str(aux_path_json), SCOPES)
                    creds = flow.run_local_server(port=0)  # Inicia o fluxo de autenticação
            else:
                # Se não houver credenciais, inicia o fluxo de autenticação
                flow = InstalledAppFlow.from_client_secrets_file(str(aux_path_json), SCOPES)
                creds = flow.run_local_server(port = 0)

            # Salva as credenciais para futuras execuções
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)

        return creds

    def generate_oauth2_string(username, access_token):
        """Gera a string de autenticação OAuth2 no formato necessário para IMAP."""
        auth_string = f'user={username}\1auth=Bearer {access_token}\1\1'
        return auth_string.encode('ascii')

    def connect_to_gmail_imap(creds, conta_email):
        """Conecta ao Gmail via IMAP usando OAuth2."""
        imap_host = 'imap.gmail.com'
        mail = imaplib.IMAP4_SSL(imap_host)
        
        # Gera a string de autenticação OAuth2
        auth_string = generate_oauth2_string(conta_email, creds.token)
        
        # Autentica no servidor IMAP usando OAuth2
        try:
            mail.authenticate('XOAUTH2', lambda x: auth_string)
            log_message(self.log_input, '\nAutenticação bem sucedida, seguindo para baixar emails\n')
        except Exception as e:
            log_message(self.log_input, f'Ocorreu um erro durante a autenticação com servidor: {e}')
            return None
        
        return mail

    # Inicie a autenticação
    creds = get_gmail_service()

    # Conecte ao Gmail via IMAP usando as credenciais OAuth2
    mail = connect_to_gmail_imap(creds, conta_email)

    if mail:
        # Selecionar a caixa de entrada ou outro rótulo
        mail.select(f'"{caixa_emails}"')

        try:
            processed_files = []

            idx = 1

            for xml in substrings_hp:
                status, email_ids = mail.search(None, f'(SUBJECT "{xml}")')
                log_message(self.log_input, f'{idx} - Buscando emails com assunto: "{xml}"')

                for email_id in email_ids[0].split():
                    status, email_data = mail.fetch(email_id, '(RFC822)')
                    raw_email = email_data[0][1]
                    message = email.message_from_bytes(raw_email)

                    # Processar o e-mail e anexos
                    for part in message.walk():
                        if part.get_content_maintype() == 'multipart':
                            continue
                        if part.get('Content-Disposition') is None:
                            continue

                        file_name = part.get_filename()

                        # Verificar se o anexo já foi processado
                        if file_name and file_name.endswith('-nfe.xml') and file_name not in processed_files:
                            log_message(self.log_input, f'{idx} - Anexo baixado: {file_name}\n')

                            try:
                                save_path = os.path.join(aux_path_XML_destino, file_name)
                                os.makedirs(aux_path_XML_destino, exist_ok = True)

                                # Salvar o arquivo
                                with open(save_path, 'wb') as f:
                                    f.write(part.get_payload(decode = True))

                                if os.path.exists(save_path):
                                    destino_path = os.path.join(aux_path_XML_destino, file_name)
                                    shutil.move(save_path, destino_path)

                                    # Adicionar o arquivo processado na lista
                                    processed_files.append(file_name)
                                else:
                                    log_message(self.log_input, f'{idx} - O arquivo de origem não foi encontrado: {save_path}')

                                    idx += 1

                            except Exception as e:
                                log_message(self.log_input, f'{idx} - Ocorreu um erro ao processar o anexo "{file_name}": {e}')

                                idx += 1
                idx += 1

        except Exception as e:
            log_message(self.log_input, f'Não foi possível prosseguir com a busca no email: {e}')
    else:
        log_message(self.log_input, f'A conexão IMAP falhou.')

    substrings_hp.clear()

def comparar_cprod(xml_file, cprod_excel, nNF_excel):
    try:
        tree = minidom.parse(xml_file)
        nf_list = tree.getElementsByTagName('nNF')  # Lista com os números de nNF
        det_list = tree.getElementsByTagName('det')  # Lista com os números de det

        # Verifica se há nNF correspondente
        for nf in nf_list:
            if nf.firstChild and nf.firstChild.data.strip() == nNF_excel.strip():  # Compara nNF
                # Para o nNF correspondente, procura os elementos <det>
                for det in det_list:
                    prod_element = det.getElementsByTagName('prod')[0] if det.getElementsByTagName('prod') else None  # Verifica se <prod> existe
                    if prod_element:
                        cProd = prod_element.getElementsByTagName('cProd')[0].firstChild.data.strip() if prod_element.getElementsByTagName('cProd') else None
                        if cProd and cProd == cprod_excel.strip():  # Compara cProd do XML com o valor do Excel
                            return det  # Retorna o <det> correspondente
        return None  # Retorna None se não houver correspondência
    except Exception as e:
        print(f"Erro ao processar o arquivo XML {xml_file}: {e}")
        return None

def formatar_planilha(worksheet):

    estilo_entrada = NamedStyle(name = 'estilo_entrada')
    estilo_entrada.font = Font(color = 'FFFFFF', name = 'consolas', size = 11)
    estilo_entrada.alignment = Alignment(horizontal = 'center', vertical = 'center')
    estilo_entrada.border = Border(left=Side(style = 'thin', color = 'FFFFFF'),
                                right=Side(style = 'thin', color = 'FFFFFF'),
                                top=Side(style = 'thin', color = 'FFFFFF'),
                                bottom=Side(style = 'thin', color = 'FFFFFF'))
    estilo_entrada.fill = PatternFill(start_color = '000000', end_color ='000000', fill_type = 'solid')
    for row in worksheet.iter_rows(min_row = 1, max_row = worksheet.max_row, min_col = 1, max_col = worksheet.max_column):
        for cell in row:  
            cell.style = estilo_entrada

def formatar_linha_difal(worksheet, cor_fundo, cor_fonte, cor_borda):
        
        for linha in worksheet.iter_rows(min_row = 1, max_row = 1, min_col = 1, max_col = 15):
            for celula in linha:
                celula.fill = PatternFill(start_color = cor_fundo, end_color = cor_fundo, fill_type = 'solid')
                celula.font = Font(color = cor_fonte, name = 'calibri', size = 10, bold = True)
                celula.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                celula.border = Border(left = Side (style = 'thin', color = cor_borda),
                                    right = Side (style = 'thin', color = cor_borda),
                                    top = Side (style = 'thin', color = cor_borda),
                                    bottom = Side (style = 'thin', color = cor_borda),
                                    )

def extrair_MO(texto):
    padrao = r'MO:([A-Za-z0-9\-]+)'

    resultado = re.search(padrao, texto)

    if resultado:
        return resultado.group(1)
    else:
        return None

# Botoes_entrada_Dell e HP
def biparxml(self, *args):

    substrings_bipar = []
    
    for index, row in planilha_df.iterrows():

        cell_bipar = row[1]

        if isinstance(cell_bipar, str) and len(cell_bipar) >= 34:
            # Extrai a substring e adiciona à lista
            substring_bipar = cell_bipar
            substrings_bipar.append(substring_bipar)
        else:
            substrings_bipar.append(cell_bipar)

    idx = 1

    # Inicializa o driver (navegador)
    driver = webdriver.Chrome()

    # Acessa a página desejada
    driver.get('https://innovation.uninet.com.br/Ulog/Main.asp?All=reset')

    time.sleep(5)

    try:
        element_usuario = WebDriverWait(driver, 500).until(
            EC.presence_of_element_located((By.NAME, 'login'))
        )
        element_usuario.send_keys(usuario_elogistica)

        element_senha = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.NAME, 'password'))
        )
        element_senha.send_keys(senha_elogistica)

        element_submit = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.NAME, 'submit'))
        )
        element_submit.click()

        time.sleep(5)

        element_botao_notificar_recebimento = WebDriverWait(driver, 500).until(
            EC.presence_of_element_located((By.XPATH, "(//input[@id='submit'])[2]"))
        )
        element_botao_notificar_recebimento.click()
    except:
        log_message(self.log_input, 'Elemento não apareceu dentro do tempo esperado.')

    time.sleep(10)

    try:
        for x in substrings_bipar:

            element_xml = WebDriverWait(driver, 500).until(
                EC.presence_of_element_located((By.NAME, 'keynfe'))
            )
            element_xml.send_keys(x)

            time.sleep(2)

            element_submit_xml = driver.find_element(By.NAME, 'submit')
            element_submit_xml.click()
            
            log_message(self.log_input, f'{idx} - Arquivo {x} processado')

            idx += 1

            time.sleep(3)
            
    except Exception as e:
        log_message(self.log_input, f'Não foi possível adicionar o arquivo {x}: {e}\n')

    time.sleep(5)

    substrings_bipar.clear()

    # Fechar o navegador
    driver.quit()

# Botoes_entrada_Dell
def criar_planilha_entrada_nf_DELL(self, *args):

    if not os.path.exists(aux_path_XML_destino):
        log_message(self.log_input, 'A pasta especificada não existe.\n')
    else:
        # Lista todos os arquivos na pasta
        arquivos = os.listdir(aux_path_XML_destino)

        # Cria uma nova planilha
        workbook = Workbook()
        sheet = workbook.active

        # Lista para armazenar os dados não formatados
        dados_nao_formatados = []

        # Itera sobre cada arquivo na pasta
        for idx, arquivo in enumerate(arquivos, start = 1):

            # Verifica se o caminho é um arquivo (e não um diretório)
            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                        
                    # Abre o arquivo XML
                    with open(caminho_arquivo, 'r') as file:
                        nfe = minidom.parse(file)

                        # Código que extrai o número do chamado
                        chamado = nfe.getElementsByTagName('infCpl')
                        print_chamado = str((chamado[0].firstChild.data[30:39]))

                        # Código que extrai o número da NF
                        ref_nf = nfe.getElementsByTagName('nNF')
                        print_ref_nf = int((ref_nf[0].firstChild.data))

                        cProds = nfe.getElementsByTagName('cProd')
                        # Iterar sobre todos os elementos 'cProd' e adicionar à lista
                        produtos = [cProd.firstChild.nodeValue for cProd in cProds]

                        # Adiciona os dados à lista de dados não formatados
                        for produto in produtos:
                            dados_nao_formatados.append((print_chamado, print_ref_nf, produto))

                except Exception as e:
                    log_message(self.log_input, f'Erro ao processar o arquivo {arquivo}: {e}\n')

        # Cria um DataFrame pandas com os dados
        df = pd.DataFrame(dados_nao_formatados, columns=['Chamado', 'NF', 'Produto'])

        # Ordena os números da coluna 'Chamado' e 'NF' do menor para o maior
        df['Chamado'] = pd.to_numeric(df['Chamado'])
        df['NF'] = pd.to_numeric(df['NF'])
        df = df.sort_values(['Chamado', 'NF'])

        # Remove as linhas duplicadas com base nos valores de 'Chamado', 'NF' e 'Produto'
        df_sem_duplicatas = df.drop_duplicates()

        # Converte o DataFrame de volta para uma lista de tuplas
        dados_sem_duplicatas = [tuple(x) for x in df_sem_duplicatas.values]

        # Insere os dados na planilha
        for dados in dados_sem_duplicatas:
            sheet.append(dados)

        # Mescla células com valores duplicados na coluna 'Chamado'
        row = 1
        while row <= sheet.max_row:
            chamado_atual = sheet.cell(row = row, column = 1).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column=1).value == chamado_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 1, end_row = row, end_column = 1)

            row += 1

        # Mescla células com valores duplicados na coluna 'NF'
        row = 1
        while row <= sheet.max_row:
            NF_atual = sheet.cell(row = row, column = 2).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column = 2).value == NF_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 2, end_row = row, end_column = 2)
                
            row += 1

        # Aplica formatação à planilha
        formatar_planilha(sheet)

        # Salva a planilha
        workbook.save('Planilha Notas de entrada_Dell.xlsx')


        if workbook.save:
            log_message(self.log_input, 'Planilha criada com sucesso: Planilha Notas de entrada_Dell.xlsx\n')

            if os.path.exists(f'{download_dir}\\Planilha Notas de entrada_Dell.xlsx'):
                os.remove(f'{download_dir}\\Planilha Notas de entrada_Dell.xlsx')

            shutil.move('Planilha Notas de entrada_Dell.xlsx', download_dir)
        else:
            log_message(self.log_input, 'Não foi possível criar a planilha\n')

# Botoes_entrada_Dell
def baixar_arquivosXML_DELL(self, *args):
    
    def mostrar_confirmacao(self):
        box = BoxLayout(orientation = 'vertical', padding = 10, spacing = 10)
        
        # Texto do aviso
        label = Label(text = f'Caso existam arquivos na pasta:\n\n {aux_path_XML_destino}, serão excluídos.\n\nDeseja continuar?', halign = 'center', valign = 'center', text_size = (400, None))
        
        # Botões "Sim" e "Não"
        botoes = BoxLayout(size_hint_y = 0.3, spacing = 10)
        botao_sim = Button(text = "Sim")
        botao_nao = Button(text = "Não")
        
        botoes.add_widget(botao_sim)
        botoes.add_widget(botao_nao)
        
        # Adiciona o texto e os botões ao layout principal do popup
        box.add_widget(label)
        box.add_widget(botoes)
        
        # Configura o popup
        popup = Popup(title = 'Aviso', content = box, size_hint = (0.5, 0.4))

        def continuar(*args):

            # Lista todos os arquivos na pasta
            arquivos_excluir = os.listdir(aux_path_XML_destino)

            for arquivo_excluir in arquivos_excluir:
                caminho_arquivo_excluir = os.path.join(aux_path_XML_destino, arquivo_excluir)
                if os.path.isfile(caminho_arquivo_excluir):
                    os.remove(caminho_arquivo_excluir)

            conectar_email_e_baixar_arquivos_Dell(self)

            popup.dismiss()

        def parar(instance):
            popup.dismiss()
            return

        botao_sim.bind(on_release = continuar)
        botao_nao.bind(on_release = parar)
    
        # Exibe o popup
        popup.open()

    mostrar_confirmacao(self)

# Botoes_entrada_Dell
def importar_produtos(self, *args):  

    idx = 1

    # Inicializa o driver (navegador)
    driver = webdriver.Chrome()

    # Acessa a página desejada
    driver.get('https://sso.iob.com.br/signin/?response_type=code&scope=&client_id=c17d4225-9d57-401b-b4fd-32503121f55b&redirect_uri=https://emissor.iob.com.br')

    try:

        element_usuario_emitir = WebDriverWait(driver, 200).until( # Busca pela página o campo de usuário
            EC.presence_of_element_located((By.ID, 'username'))
        )
        element_usuario_emitir.send_keys(usuario_IOB) # Insere o nome de usuário

        time.sleep(2.5)

        element_senha_emitir = WebDriverWait(driver, 50).until( # Busca pela página o campo de senha
            EC.presence_of_element_located((By.ID, 'password'))
        )
        element_senha_emitir.send_keys(senha_IOB) # Insere a senha

        time.sleep(2.5) # tempo para carregar o reCAPTCHA

        element_reCAPTCHA_emitir = WebDriverWait(driver, 50).until( # Buscar o elemento reCAPTCHA na página
            EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/div[2]/div/div[2]/main/div[2]/div/form/div[4]/div/div/div/iframe'))
        )
        driver.switch_to.frame(element_reCAPTCHA_emitir) # Muda para o contexto do iframe onde o captcha está inserido
        captcha_checkbox = driver.find_element(By.XPATH, '//div[@class="recaptcha-checkbox-border"]') # Busca o elemento clicável do captcha
        captcha_checkbox.click() # Clica no elemento captcha

        driver.switch_to.default_content() # Muda para o contexto padrão da página

        time.sleep(10)

        element_submit_emitir = WebDriverWait(driver, 50).until( # Busca pelo elemento submit na página
            EC.presence_of_element_located((By.XPATH, '//*[@id="formButton"]'))
        )
        element_submit_emitir.click()  # Clica no elemento submit

        element_emissão_notas_emitir = WebDriverWait(driver, 200).until( # Busca pelo botao de emissão de notas
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div[title="Emissão de Notas"]'))
        )
                
        if element_emissão_notas_emitir:
            driver.get('https://emissor2.iob.com.br/notafiscal/incoming_nfes') # Caso encontre o botão de notas, entra na página para emitir

        try:
            for x in os.listdir(aux_path_XML_destino):
                # Remover o overlay antes de qualquer clique
                try:
                    driver.execute_script("""
                        let overlay = document.getElementById('UIDialogLayer');
                        if (overlay) {
                            overlay.remove();  // Remove completamente o overlay da página
                        }
                    """)
                    time.sleep(0.5)  # Pequena pausa para garantir que o overlay seja removido
                except Exception as e:
                    print("Não foi possível remover o overlay:", e)

                # Iniciar o processo de importação
                element_importar = WebDriverWait(driver, 200).until(
                    EC.presence_of_element_located((By.XPATH, '//*[contains(@class, "UILink button pull-right")]'))
                )
                time.sleep(0.5)
                element_importar.click()

                # Aguardar e clicar no input de arquivo
                element_abrir_input = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="import_nfe_nfe_xml"]'))
                )
                caminho = os.path.join(aux_path_XML_destino, x)
                element_abrir_input.send_keys(caminho)

                # Aguardar o botão de "Confirmar" (ou similar) aparecer após o upload
                element_confirmar = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[contains(@class, "save button primary")]'))
                )
                element_confirmar.click()

                # Navegar de volta para a página inicial (após enviar um arquivo)
                driver.get('https://emissor2.iob.com.br/notafiscal/incoming_nfes')

                # Aguardar a página carregar e clicar novamente no botão "Importar"
                element_importar1 = WebDriverWait(driver, 200).until(
                    EC.presence_of_element_located((By.XPATH, '//*[contains(@class, "UILink button pull-right")]'))
                )
                time.sleep(0.5)
                element_importar1.click()

                log_message(self.log_input, f'{idx}: {x} importado com sucesso.')

                idx += 1

                # Aguardar um breve momento para garantir que a próxima iteração não aconteça muito rápido
                time.sleep(2)

        except Exception as e:
            log_message(self.log_input, f'Erro encontrado: {e}')

    except Exception as e:
        log_message(self.log_input, f'Ocorreu um erro: {e}')

# Botoes_devolução_Dell
def valores_devolução_DELL(self, *args):

    def mostrar_confirmacao(self):
        box = BoxLayout(orientation = 'vertical', padding = 10, spacing = 10)
        
        # Texto do aviso
        label = Label(text = f'Caso existam arquivos na pasta:\n\n {aux_path_XML_destino}, serão excluídos.\n\nDeseja continuar?', halign = 'center', valign = 'center', text_size = (400, None))
        
        # Botões "Sim" e "Não"
        botoes = BoxLayout(size_hint_y = 0.3, spacing = 10)
        botao_sim = Button(text = "Sim")
        botao_nao = Button(text = "Não")
        
        botoes.add_widget(botao_sim)
        botoes.add_widget(botao_nao)
        
        # Adiciona o texto e os botões ao layout principal do popup
        box.add_widget(label)
        box.add_widget(botoes)
        
        # Configura o popup
        popup = Popup(title = 'Aviso', content = box, size_hint = (0.5, 0.4))
        
        # Função para o botão "Sim"
        def continuar(instance):

            # Lista todos os arquivos na pasta
            arquivos_excluir = os.listdir(aux_path_XML_destino)

            for arquivo_excluir in arquivos_excluir:
                caminho_arquivo_excluir = os.path.join(aux_path_XML_destino, arquivo_excluir)
                if os.path.isfile(caminho_arquivo_excluir):
                    os.remove(caminho_arquivo_excluir)

            conectar_email_e_baixar_arquivos_Dell(self)
            
            popup.dismiss()

            idx_planilha = 1

            # Inicializa o driver (navegador)
            driver = webdriver.Chrome()

            # Acessa a página desejada
            driver.get('https://innovation.uninet.com.br/Ulog/Main.asp?All=reset')

            try:
                element_usuario = WebDriverWait(driver, 200).until(
                    EC.presence_of_element_located((By.NAME, 'login'))
                )
                element_usuario.send_keys(usuario_elogistica)

                element_senha = driver.find_element(By.NAME, 'password')
                element_senha.send_keys(senha_elogistica)

                element_submit = driver.find_element(By.NAME, 'submit')
                element_submit.click()

                time.sleep(5)

                element_peças_pendentes_devolução = WebDriverWait(driver, 200).until(
                    EC.presence_of_element_located((By.XPATH, '(//input[@id="submit"])[4]'))
                )
                element_peças_pendentes_devolução.click()
            except:
                log_message(self.log_input, 'O site demorou mais que o esperado para carregar.')
                
            resultados = {}
            
            for index, row in planilha_df.iloc[:].iterrows():
                CHAMADO = str(row[0]).strip()
                NF = str(row[1]).strip()
                PN = str(row[2]).strip()
                PPID = str(row[4]).strip()
                STATUS_planilha = str(row[9]).strip()

                if STATUS_planilha.upper() == 'DEFECTIVE' or STATUS_planilha.upper() == 'DOA':

                    if PPID.upper() == 'X':
                        log_message(self.log_input, f'{idx_planilha}: Peça sem PPID na planilha.\n')

                        try:
                                    # Encontra o valor do produto
                                    element_Vprod = elements.find_element(By.XPATH, './/td[9]')
                                    element_Vprod_value = element_Vprod.text

                                    # Adiciona o valor ao dicionário
                                    resultados[f"{PN}_{STATUS_planilha}"] = {'VALOR': element_Vprod_value, 'CST': None, 'NCM': None, 'QTD.': None}

                                    if not os.path.exists(aux_path_XML_destino):
                                        continue
                                    else:  
                                        # Lista todos os arquivos na pasta
                                        arquivos = os.listdir(aux_path_XML_destino)

                                        for arquivo in arquivos:
                                            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)

                                            if os.path.isfile(caminho_arquivo):
                                                try:
                                                    # Abrir o arquivo XML no modo leitura
                                                    with open(caminho_arquivo, 'r'):

                                                        # Comparar cProd do arquivo XML com o valor PN do Excel
                                                        det_element = comparar_cprod(caminho_arquivo, PN, NF)

                                                        if det_element:
                                                            # Buscar o valor de <orig> e <NCM> dentro do mesmo <det>
                                                            CST = det_element.getElementsByTagName('orig')
                                                            NCM_prod = det_element.getElementsByTagName('NCM')
                                                            Quantidade_prod = det_element.getElementsByTagName('qCom')
                                                            
                                                            if CST and NCM_prod:
                                                                resultados[f'{PN}_{STATUS_planilha}']['CST'] = str(CST[0].firstChild.data)
                                                                resultados[f'{PN}_{STATUS_planilha}']['NCM'] = str(NCM_prod[0].firstChild.data)
                                                                resultados[f'{PN}_{STATUS_planilha}']['QTD.'] = float(Quantidade_prod[0].firstChild.data)

                                                            else:
                                                                log_message(self.log_input, f'Elementos CST, NCM e/ou Quantidade não encontrados no arquivo {arquivo}.')
                                                except Exception as e:
                                                    log_message(self.log_input, f'Erro ao processar o arquivo {arquivo}: {e}\n')
                                    idx_planilha += 1

                        except Exception as e:
                            log_message(self.log_input, f'{idx_planilha}: Chamado {CHAMADO}/{PN} não consta no site.\n')
                            idx_planilha += 1
                            continue

                    else:
                        try:
                            
                            # Tenta encontrar o elemento correspondente
                            elements = WebDriverWait(driver, 20).until(
                                EC.presence_of_element_located((By.XPATH, f'//tr[contains(td[4], "{CHAMADO}") and contains(td[6], "{PN}")]'))
                            )

                            time.sleep(1)

                            if elements:
                                try:
                                    # Encontra o valor do produto
                                    element_Vprod = elements.find_element(By.XPATH, './/td[9]')
                                    element_Vprod_value = element_Vprod.text

                                    # Adiciona o valor ao dicionário
                                    resultados[f"{PN}_{STATUS_planilha}"] = {'VALOR': element_Vprod_value, 'CST': None, 'NCM': None, 'QTD.': None}

                                    time.sleep(1)

                                    # Encontra o link <a> na mesma linha
                                    element_link = elements.find_element(By.XPATH, './/td[5]')
                                    element_link.click()

                                    time.sleep(1)
                                    
                                    # Insere o PPID da planilha no site
                                    inserir_ppid = WebDriverWait(driver, 60).until(
                                    EC.presence_of_element_located((By.NAME, 'PPID'))
                                    )
                                    inserir_ppid.send_keys(PPID)

                                    botao_enviar_ppid = WebDriverWait(driver, 20).until(
                                    EC.presence_of_element_located((By.NAME, 'submit'))
                                    )
                                    botao_enviar_ppid.click()

                                    time.sleep(1)

                                    alerta = driver.switch_to.alert
                                    alerta.accept()

                                    if not os.path.exists(aux_path_XML_destino):
                                        continue
                                    else:  
                                        # Lista todos os arquivos na pasta
                                        arquivos = os.listdir(aux_path_XML_destino)

                                        for arquivo in arquivos:
                                            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)

                                            if os.path.isfile(caminho_arquivo):
                                                try:
                                                    # Abrir o arquivo XML no modo leitura
                                                    with open(caminho_arquivo, 'r'):

                                                        # Comparar cProd do arquivo XML com o valor PN do Excel
                                                        det_element = comparar_cprod(caminho_arquivo, PN, NF)

                                                        if det_element:
                                                            # Buscar o valor de <orig> e <NCM> dentro do mesmo <det>
                                                            CST = det_element.getElementsByTagName('orig')
                                                            NCM_prod = det_element.getElementsByTagName('NCM')
                                                            Quantidade_prod = det_element.getElementsByTagName('qCom')
                                                            
                                                            if CST and NCM_prod:
                                                                resultados[f'{PN}_{STATUS_planilha}']['CST'] = str(CST[0].firstChild.data)
                                                                resultados[f'{PN}_{STATUS_planilha}']['NCM'] = str(NCM_prod[0].firstChild.data)
                                                                resultados[f'{PN}_{STATUS_planilha}']['QTD.'] = float(Quantidade_prod[0].firstChild.data)

                                                            else:
                                                                log_message(self.log_input, f'Elementos CST ou NCM não encontrados no arquivo {arquivo}.')
                                                except Exception as e:
                                                    log_message(self.log_input, f'Erro ao processar o arquivo {arquivo}: {e}\n')

                                            element_peças_pendentes_devolução1 = WebDriverWait(driver, 200).until(
                                            EC.presence_of_element_located((By.XPATH, '(//input[@id="submit"])[4]'))
                                            )

                                    element_peças_pendentes_devolução1.click()
            
                                except Exception as e:
                                    print()

                                log_message(self.log_input, f'{idx_planilha}: Chamado {CHAMADO}/{PN} inserido com sucesso.\n')
                                idx_planilha += 1

                        except Exception as e:

                            log_message(self.log_input, f'{idx_planilha}: Chamado {CHAMADO}/{PN} não consta no site.\n')
                            idx_planilha += 1


                elif STATUS_planilha == 'GOOD':
                    try:
                        if os.path.exists(aux_path_XML_destino):
                            arquivos = os.listdir(aux_path_XML_destino)

                            for arquivo in arquivos:
                                caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)

                                if os.path.isfile(caminho_arquivo):
                                    det_element = comparar_cprod(caminho_arquivo, PN, NF)

                                    if det_element:
                                        vProd = det_element.getElementsByTagName('vProd')
                                        CST = det_element.getElementsByTagName('orig')
                                        NCM_prod = det_element.getElementsByTagName('NCM')
                                        Quantidade_prod = det_element.getElementsByTagName('qCom')
                                        
                                        if CST and NCM_prod and vProd:
                                            # Adiciona as informações ao dicionário
                                            resultados[f"{PN}_{STATUS_planilha}"] = {
                                            'VALOR': str(vProd[0].firstChild.data) if vProd else None,
                                            'CST': str(CST[0].firstChild.data) if CST else None,
                                            'NCM': str(NCM_prod[0].firstChild.data) if NCM_prod else None,
                                            'QTD.': float(Quantidade_prod[0].firstChild.data) if Quantidade_prod else None
                                        }
                                        else:
                                            log_message(self.log_input, f'Elementos CST ou NCM não encontrados no arquivo {arquivo}.')
                                            
                        log_message(self.log_input, f'{idx_planilha}: Chamado {CHAMADO}/{PN} inserido com sucesso.\n')
                        idx_planilha += 1
                        
                    except Exception as e:
                        log_message(self.log_input, f'Não foi possível abrir o arquivo {idx_planilha}: {e}')

                        idx_planilha += 1

            # Criando uma cópia das colunas selecionadas do DataFrame original
            planilha_copia = planilha_df.iloc[:, [0, 1, 2]].copy()

            # Renomeando as colunas da cópia
            planilha_copia.columns = ['CHAMADO', 'REF NF', 'PART NUMBER']

            # Adicionando as colunas ao DataFrame a partir do dicionário
            planilha_copia['VALOR'] = [
                resultados.get(f"{row[2]}_{row[9]}", {}).get('VALOR', None)
                for _, row in planilha_df.iterrows()
            ]

            planilha_copia['CST'] = [
                resultados.get(f"{row[2]}_{row[9]}", {}).get('CST', None)
                for _, row in planilha_df.iterrows()
            ]

            planilha_copia['NCM'] = [
                resultados.get(f"{row[2]}_{row[9]}", {}).get('NCM', None)
                for _, row in planilha_df.iterrows()
            ]

            planilha_copia['QTD.'] = [
                resultados.get(f'{row[2]}_{row[9]}', {}).get('QTD.', None)
                for _, row in planilha_df.iterrows()
            ]

            # Substituindo pontos por vírgulas na coluna VALOR
            planilha_copia['VALOR'] = planilha_copia['VALOR'].astype(str).str.replace('.', ',', regex = False)

            # Exportando o DataFrame para Excel
            planilha_copia.to_excel('Planilha Dell valores_devolução formatada.xlsx', index = False)

            if os.path.exists(f'{download_dir}\\Planilha Dell valores_devolução formatada.xlsx'):
                os.remove(f'{download_dir}\\Planilha Dell valores_devolução formatada.xlsx')

            shutil.move('Planilha Dell valores_devolução formatada.xlsx', download_dir)

            log_message(self.log_input, f'Planilha criada com sucesso: Planilha Dell valores_devolução formatada.xlsx')
            
            time.sleep(5)
            driver.quit()

        def parar(instance):
            popup.dismiss()
            return

        botao_sim.bind(on_release = continuar)
        botao_nao.bind(on_release = parar)
    
        # Exibe o popup
        popup.open()

    mostrar_confirmacao(self)

# Botoes_entrada_HP
def baixar_arquivosXML_HP(self, *args):
        
    def mostrar_confirmacao(self):
        box = BoxLayout(orientation = 'vertical', padding = 10, spacing = 10)
        
        # Texto do aviso
        label = Label(text = f'Caso existam arquivos na pasta:\n\n {aux_path_XML_destino}, serão excluídos.\n\nDeseja continuar?', halign = 'center', valign = 'center', text_size = (400, None))
        
        # Botões "Sim" e "Não"
        botoes = BoxLayout(size_hint_y = 0.3, spacing = 10)
        botao_sim = Button(text = "Sim")
        botao_nao = Button(text = "Não")
        
        botoes.add_widget(botao_sim)
        botoes.add_widget(botao_nao)
        
        # Adiciona o texto e os botões ao layout principal do popup
        box.add_widget(label)
        box.add_widget(botoes)
        
        # Configura o popup
        popup = Popup(title = 'Aviso', content = box, size_hint = (0.5, 0.4))

        def continuar(*args):

            # Lista todos os arquivos na pasta
            arquivos_excluir = os.listdir(aux_path_XML_destino)

            for arquivo_excluir in arquivos_excluir:
                caminho_arquivo_excluir = os.path.join(aux_path_XML_destino, arquivo_excluir)
                if os.path.isfile(caminho_arquivo_excluir):
                    os.remove(caminho_arquivo_excluir)

            conectar_email_e_baixar_arquivos_HP(self)

            popup.dismiss()

        def parar(instance):
            popup.dismiss()
            return

        botao_sim.bind(on_release = continuar)
        botao_nao.bind(on_release = parar)
    
        # Exibe o popup
        popup.open()

    mostrar_confirmacao(self)

# Botoes_entrada_HP
def criar_planilha_entrada_nf_HP(self, *args):

    if not os.path.exists(aux_path_XML_destino):
        log_message(self.log_input, 'A pasta especificada não existe.\n')
    else:
        # Lista todos os arquivos na pasta
        arquivos = os.listdir(aux_path_XML_destino)

        # Cria uma nova planilha
        workbook = Workbook()
        sheet = workbook.active

        # Lista para armazenar os dados não formatados
        dados_nao_formatados = []

        # Itera sobre cada arquivo na pasta
        for idx, arquivo in enumerate(arquivos, start = 1):

            # Verifica se o caminho é um arquivo (e não um diretório)
            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                        
                    # Abre o arquivo XML
                    with open(caminho_arquivo, 'r') as file:
                        nfe = minidom.parse(file)

                        # Código que extrai o número do chamado
                        chamado = nfe.getElementsByTagName('infCpl')
                        if chamado:
                            print_chamado = chamado[0].firstChild.data
                            mo_valor = extrair_MO(print_chamado)
                        else:
                            mo_valor = 'MO não encontrado'

                        # Código que extrai o número da NF
                        ref_nf = nfe.getElementsByTagName('nNF')
                        print_ref_nf = int((ref_nf[0].firstChild.data))

                        cProds = nfe.getElementsByTagName('cProd')
                        # Iterar sobre todos os elementos 'cProd' e adicionar à lista
                        produtos = [cProd.firstChild.nodeValue for cProd in cProds]

                        # Adiciona os dados à lista de dados não formatados
                        for produto in produtos:
                            dados_nao_formatados.append((mo_valor, print_ref_nf, produto))

                except Exception as e:
                    log_message(self.log_input, f'Erro ao processar o arquivo {arquivo}: {e}\n')

        # Cria um DataFrame pandas com os dados
        df = pd.DataFrame(dados_nao_formatados, columns=['Chamado', 'NF', 'Produto'])

        # Ordena os números da coluna 'Chamado' e 'NF' do menor para o maior
        df['Chamado'] = (df['Chamado'])
        df['NF'] = pd.to_numeric(df['NF'])
        df = df.sort_values(['Chamado', 'NF'])

        # Remove as linhas duplicadas com base nos valores de 'Chamado', 'NF' e 'Produto'
        df_sem_duplicatas = df.drop_duplicates()

        # Converte o DataFrame de volta para uma lista de tuplas
        dados_sem_duplicatas = [tuple(x) for x in df_sem_duplicatas.values]

        # Insere os dados na planilha
        for dados in dados_sem_duplicatas:
            sheet.append(dados)

        # Mescla células com valores duplicados na coluna 'Chamado'
        row = 1
        while row <= sheet.max_row:
            chamado_atual = sheet.cell(row = row, column = 1).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column=1).value == chamado_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 1, end_row = row, end_column = 1)

            row += 1

        # Mescla células com valores duplicados na coluna 'NF'
        row = 1
        while row <= sheet.max_row:
            NF_atual = sheet.cell(row = row, column = 2).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column = 2).value == NF_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 2, end_row = row, end_column = 2)
                
            row += 1

        # Aplica formatação à planilha
        formatar_planilha(sheet)

        # Salva a planilha
        workbook.save('Planilha Notas de entrada_HP.xlsx')

        if workbook.save:

            log_message(self.log_input, 'Planilha criada com sucesso: Planilha Notas de entrada_HP\n')

            if os.path.exists(f'{download_dir}\\Planilha Notas de entrada_HP.xlsx'):
                os.remove(f'{download_dir}\\Planilha Notas de entrada_HP.xlsx')

            shutil.move('Planilha Notas de entrada_HP.xlsx', download_dir)

        else:
            log_message(self.log_input, 'Não foi possível criar a planilha\n')

# Botoes_devolução_HP
def valores_devolução_HP(self, *args):

    def mostrar_confirmacao(self):
        box = BoxLayout(orientation = 'vertical', padding = 10, spacing = 10)
        
        # Texto do aviso
        label = Label(text = f'Caso existam arquivos na pasta:\n\n {aux_path_XML_destino}, serão excluídos.\n\nDeseja continuar?', halign = 'center', valign = 'center', text_size = (400, None))
        
        # Botões "Sim" e "Não"
        botoes = BoxLayout(size_hint_y = 0.3, spacing = 10)
        botao_sim = Button(text = "Sim")
        botao_nao = Button(text = "Não")
        
        botoes.add_widget(botao_sim)
        botoes.add_widget(botao_nao)
        
        # Adiciona o texto e os botões ao layout principal do popup
        box.add_widget(label)
        box.add_widget(botoes)
        
        # Configura o popup
        popup = Popup(title = 'Aviso', content = box, size_hint = (0.5, 0.4))
        
        # Função para o botão "Sim"
        def continuar(instance):

            # Lista todos os arquivos na pasta
            arquivos_excluir = os.listdir(aux_path_XML_destino)

            for arquivo_excluir in arquivos_excluir:
                caminho_arquivo_excluir = os.path.join(aux_path_XML_destino, arquivo_excluir)
                if os.path.isfile(caminho_arquivo_excluir):
                    os.remove(caminho_arquivo_excluir)

            conectar_email_e_baixar_arquivos_HP(self)
            
            popup.dismiss()

            idx_planilha = 1

            resultados_hp = {}

            for index, row in planilha_df.iloc[:].iterrows():
                CHAMADO = str(row[0]).strip()
                NF = str(row[1]).strip()
                PN = str(row[2]).strip()
                STATUS_planilha = str(row[9]).strip()

                try:
                    if os.path.exists(aux_path_XML_destino):

                        arquivos = os.listdir(aux_path_XML_destino)

                        for arquivo in arquivos:
                            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)

                            if os.path.isfile(caminho_arquivo):
                                det_element = comparar_cprod(caminho_arquivo, PN, NF)

                                if det_element:
                                        
                                    vProd = det_element.getElementsByTagName('vProd')

                                    CST = det_element.getElementsByTagName('orig')

                                    NCM_prod = det_element.getElementsByTagName('NCM')

                                    Quantidade_prod = det_element.getElementsByTagName('qCom')
                                    
                                    if CST and NCM_prod and vProd:

                                        if STATUS_planilha.upper() == 'GOOD':

                                            resultados_hp[f"{PN}_{STATUS_planilha}"] = {
                                            'VALOR': float(f'{float(vProd[0].firstChild.data):.2f}') if vProd else None,
                                            'CST': str(CST[0].firstChild.data) if CST else None,
                                            'NCM': str(NCM_prod[0].firstChild.data) if NCM_prod else None,
                                            'QTD.': float(Quantidade_prod[0].firstChild.data) if Quantidade_prod else None
                                        }
                                            
                                        elif STATUS_planilha in ['DEFECTIVE', 'DOA']:
                                    
                                            resultados_hp[f"{PN}_{STATUS_planilha}"] = {
                                            'VALOR': float(f'{float(vProd[0].firstChild.data) / 10:.2f}') if vProd else None,
                                            'CST': str(CST[0].firstChild.data) if CST else None,
                                            'NCM': str(NCM_prod[0].firstChild.data) if NCM_prod else None,
                                            'QTD.': float(Quantidade_prod[0].firstChild.data) if Quantidade_prod else None
                                        }
                                    else:
                                        log_message(self.log_input, f'Elementos CST ou NCM não encontrados no arquivo {arquivo}.')
                                        
                    log_message(self.log_input, f'{idx_planilha}: Chamado {CHAMADO}/{PN} inserido com sucesso.\n')
                    idx_planilha += 1
                    
                except Exception as e:
                    log_message(self.log_input, f'Não foi possível abrir o arquivo {idx_planilha}: {e}')

                    idx_planilha += 1
            
            # Criando uma cópia das colunas selecionadas do DataFrame original
            planilha_copia = planilha_df.iloc[:, [0, 1, 2]].copy()

            # Renomeando as colunas da cópia
            planilha_copia.columns = ['CHAMADO', 'REF NF', 'PART NUMBER']

            # Adicionando as colunas ao DataFrame a partir do dicionário
            planilha_copia['VALOR'] = [
                resultados_hp.get(f"{row[2]}_{row[9]}", {}).get('VALOR', None)
                for _, row in planilha_df.iterrows()
            ]

            planilha_copia['CST'] = [
                resultados_hp.get(f"{row[2]}_{row[9]}", {}).get('CST', None)
                for _, row in planilha_df.iterrows()
            ]

            planilha_copia['NCM'] = [
                resultados_hp.get(f"{row[2]}_{row[9]}", {}).get('NCM', None)
                for _, row in planilha_df.iterrows()
            ]

            planilha_copia['QTD.'] = [
                resultados_hp.get(f'{row[2]}_{row[9]}', {}).get('QTD.', None)
                for _, row in planilha_df.iterrows()
            ]

            # Substituindo pontos por vírgulas na coluna VALOR
            planilha_copia['VALOR'] = planilha_copia['VALOR'].astype(str).str.replace('.', ',', regex = False)


            # Exportando o DataFrame para Excel
            planilha_copia.to_excel('Planilha HP valores_devolução formatada.xlsx', index = False)

            if os.path.exists(f'{download_dir}\\Planilha HP valores_devolução formatada.xlsx'):
                os.remove(f'{download_dir}\\Planilha HP valores_devolução formatada.xlsx')

            shutil.move('Planilha HP valores_devolução formatada.xlsx', download_dir)

            log_message(self.log_input, f'Planilha criada com sucesso: Planilha HP valores_devolução formatada.xlsx')
            
        def parar(instance):
            popup.dismiss()
            return

        botao_sim.bind(on_release = continuar)
        botao_nao.bind(on_release = parar)
    
        # Exibe o popup
        popup.open()

    mostrar_confirmacao(self)

# Botoes_difal
def criar_planilha_difal(self, *args):

    if not os.path.exists(aux_path_XML_destino):
        log_message(self.log_input, 'A pasta especificada não existe.\n')
    else:
        # Lista todos os arquivos na pasta
        arquivos = os.listdir(aux_path_XML_destino)

        #Cria um novo objeto
        workbook = Workbook()

        sheet = workbook.active

        #Adiciona dados à planilha
        Dados_linha1 = [
        'Nome Credenciada / Destino',
        'CNPJ Credenciada / Destino',
        'Mês de Referência',
        'Número da ND',
        'CNPJ Origem',
        'Razão Social / Origem',
        'Nr. Chamado do Cliente',
        'Número da NF de referência',
        'Data de Emissão',
        'Valor da BC do ICMS da NF de referência',
        'Alíquota Interestadual',
        'Valor do ICMS da NF de referência',
        'Alíquota interna do Estado de destino',
        'Alíquota do Diferencial',
        'Valor do ICMS recolhido por Diferencial de Alíquota'
        ]
        for col, valor in enumerate(Dados_linha1, start = 1):
            sheet.cell(row = 1, column = col, value = valor)

            #Adiciona a largura e altura das colunas
            largura_colunas = [
            25.43, 18.71, 10.00, 9.86, 20.14, 29.00, 18.43, 14.00, 12.14, 23.71, 16.86, 20.57, 15.00, 15.00, 22.57
            ]

        for i, largura in enumerate(largura_colunas, start = 1):
            sheet.column_dimensions[chr(64 + i)].width = largura
                                
        for a in range(2):
            sheet.row_dimensions[a].height = 34.00

        formatar_linha_difal(sheet, cor_fundo = '9BC2E6', cor_fonte = '000000', cor_borda = '000000')

        for idx, arquivo in enumerate(arquivos, start = 2):

            if arquivo.endswith(('procNFe.xml', 'nfe.xml')):

                try:
                    # Verifica se o caminho é um arquivo (e não um diretório)
                    caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)
                    if os.path.isfile(caminho_arquivo):

                        # Abre o arquivo XML
                        with open(caminho_arquivo, 'r') as file:
                            nfe = minidom.parse(file)

                            if arquivo.endswith('procNFe.xml'):
                                # Código que extrai o número do chamado
                                chamado = nfe.getElementsByTagName('infCpl')
                                print_chamado = str((chamado[0].firstChild.data[30:39]))
                                                        
                                # Código que extrai a data de emissão
                                data_emissao = nfe.getElementsByTagName('dhEmi')
                                print_data_emissao = str((data_emissao[0].firstChild.data[2:10]))
                                data = datetime.strptime(print_data_emissao, '%y-%m-%d')
                                nova_data = data.strftime('%d/%m/%y')
                                                        
                                # Código que extrai o ICMS da nota
                                ICMS_prod = nfe.getElementsByTagName('pICMS')
                                print_ICMS_prod = float(ICMS_prod[0].firstChild.data) / 100

                                # Código que extrai a nota fiscal da nota
                                ref_nf = nfe.getElementsByTagName('nNF')
                                print_ref_nf = str(ref_nf[0].firstChild.data)

                                # Código que extrai o valor da nota
                                vNF = nfe.getElementsByTagName('vNF')
                                print_valor_vNF = float(vNF[0].firstChild.data)

                                # Código para formatar o CNPJ
                                CNPJ_dell = nfe.getElementsByTagName('CNPJ')
                                print_CNPJ_dell = str(CNPJ_dell[0].firstChild.data)

                                print_CNPJ_dell_formatado = f"{print_CNPJ_dell[:2]}.{print_CNPJ_dell[2:5]}.{print_CNPJ_dell[5:8]}/{print_CNPJ_dell[8:12]}-{print_CNPJ_dell[12:]}"

                                # Código para formatar a razão social
                                razao_social_dell = nfe.getElementsByTagName('xNome')
                                print_razao_social_dell = str(razao_social_dell[0].firstChild.data)

                            elif arquivo.endswith('nfe.xml'):
                                try:
                                    chamado = nfe.getElementsByTagName('infCpl')
                                    if chamado:
                                        print_chamado = chamado[0].firstChild.data
                                        mo_valor = extrair_MO(print_chamado)

                                    else:
                                        mo_valor = 'MO não encontrado'
                    
                                except Exception as e:
                                    log_message(self.log_input, f'MO não contrado no arquivo: {arquivo}\n')
                                    continue

                                # Código que extrai a data de emissão
                                data_emissao = nfe.getElementsByTagName('dhEmi')
                                print_data_emissao = str((data_emissao[0].firstChild.data[2:10]))
                                data = datetime.strptime(print_data_emissao, '%y-%m-%d')
                                nova_data = data.strftime('%d/%m/%y')
                                            
                                # Código que extrai o ICMS da nota
                                ICMS_prod = nfe.getElementsByTagName('pICMS')
                                print_ICMS_prod = float(ICMS_prod[0].firstChild.data) / 100

                                # Código que extrai a nota fiscal da nota
                                ref_nf = nfe.getElementsByTagName('nNF')
                                print_ref_nf = str(ref_nf[0].firstChild.data)

                                # Código que extrai o valor da nota
                                vNF = nfe.getElementsByTagName('vNF')
                                print_valor_vNF = float(vNF[0].firstChild.data)
                                            
                                # Código para formatar o CNPJ
                                CNPJ_HP = nfe.getElementsByTagName('CNPJ')              
                                print_CNPJ_HP = str(CNPJ_HP[0].firstChild.data)

                                print_CNPJ_HP_formatado = f"{print_CNPJ_HP[:2]}.{print_CNPJ_HP[2:5]}.{print_CNPJ_HP[5:8]}/{print_CNPJ_HP[8:12]}-{print_CNPJ_HP[12:]}"

                                # Código para formatar a razão social
                                razao_social_hp = nfe.getElementsByTagName('xFant')
                                print_razao_social_hp = str(razao_social_hp[0].firstChild.data)

                            configs_difal = obter_configs()

                            if configs_difal:
                                aliquota_interna = configs_difal.get('aliquota_interna')
                                nome_credenciada = configs_difal.get('nome_credenciada')
                                cnpj_credenciada = configs_difal.get('cnpj_credenciada')
                            else:
                                aliquota_interna = 'Não há informações da credenciada'
                                nome_credenciada = 'Não há informações da credenciada'
                                cnpj_credenciada = 'Não há informações da credenciada'
                                
                            # Código para formatar a coluna M da alíquota interna do estado (altera de estado para estado)
                            print_aliquota_interna = float(aliquota_interna) / 100

                            # Insere os dados na planilha
                            sheet[f'A{idx}'] = nome_credenciada
                            sheet[f'B{idx}'] = cnpj_credenciada
                                                        
                            if arquivo.endswith('procNFe.xml'):
                                sheet[f'E{idx}'] = print_CNPJ_dell_formatado
                                sheet[f'F{idx}'] = print_razao_social_dell
                                sheet[f'G{idx}'] = print_chamado 
                                 
                            elif arquivo.endswith('-nfe.xml'):
                                sheet[f'E{idx}'] = print_CNPJ_HP_formatado
                                sheet[f'F{idx}'] = print_razao_social_hp
                                sheet[f'G{idx}'] = mo_valor

                            sheet[f'H{idx}'] = print_ref_nf
                            sheet[f'I{idx}'] = nova_data
                            sheet[f'J{idx}'] = print_valor_vNF
                            sheet[f'K{idx}'] = print_ICMS_prod
                            sheet[f'L{idx}'] = print_valor_vNF * print_ICMS_prod
                            sheet[f'M{idx}'] = print_aliquota_interna
                            sheet[f'N{idx}'] = print_aliquota_interna - print_ICMS_prod
           
                except Exception as e:
                    log_message(self.log_input, f'Não foi possivel processar arquivos: {e}')

        dados = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row = 2, values_only = True):
            dados.append(row)

        dados_sorted = sorted(dados, key=lambda x: x[7])

        for row in sheet.iter_rows(min_row = 2, max_col = sheet.max_column, max_row = sheet.max_row):
            for cell in row:
                cell.value = None

        for col_index, header in enumerate(headers):
            sheet.cell(row = 1, column = col_index + 1, value = header)

        for index, row in enumerate(dados_sorted):
            for col_index, value in enumerate(row):
                sheet.cell(row = index + 2, column = col_index + 1, value = value)
        
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 10, max_col = 10):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 11, max_col = 11):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00 

        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 12, max_col = 12):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE 

        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 13, max_col = 14):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00     

        for linha in range(2, sheet.max_row + 1):
            for col in range(1, 16):
                celula = sheet.cell(row = linha, column = col)
                celula.fill = PatternFill(start_color = 'FFFFFF', end_color = 'FFFFFF', fill_type = 'solid')
                celula.font = Font(color = '000000', name = 'calibri', size = 11, bold = False)
                celula.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = False)
                celula.border = Border(left=Side(style = 'thin', color = '000000'),
                                    right=Side(style = 'thin', color = '000000'),
                                        top=Side(style = 'thin', color = '000000'),
                                        bottom=Side(style = 'thin', color = '000000'))
                
        workbook.save('Planilha Difal.xlsx')

        if workbook.save:
            log_message(self.log_input, '\nPlanilha criada com sucesso: Planilha Difal\n')

            if os.path.exists(f'{download_dir}\\Planilha Difal.xlsx'):
                os.remove(f'{download_dir}\\Planilha Difal.xlsx')

            shutil.move('Planilha Difal.xlsx', download_dir)
        else:
            log_message(self.log_input, 'Não foi possível criar a planilha\n')
         
def obter_usuarios_validos():
    """Obtém a lista de usuários válidos do servidor."""
    try:
        # Fazendo a requisição para obter a lista de usuários válidos
        response = requests.get("https://servidor-para-emme2.onrender.com/usuarios")
        
        # Verificando se a resposta foi bem-sucedida
        if response.status_code == 200:
            data = response.json()  # Converte a resposta para JSON

            if isinstance(data, dict) and "usuarios" in data:
                usuarios = data["usuarios"]
                return usuarios  # Retorna a lista de usuários
            
            else:
                return []  # Se o formato for inesperado, retorna lista vazia
        else:
            return []  # Se houver erro de status, retorna lista vazia
        
    except Exception as e:
        return []  # Se ocorrer erro na requisição, retorna lista vazia

lista_autenticar = obter_usuarios_validos()

def validar_usuario(self, instance):
    """Valida o e-mail do usuário."""    
    email = self.email_autenticar.text.strip().lower()  # Remove espaços e transforma para minúsculas
    if not email:
        self.label_email_autenticar.text = "Por favor, insira um e-mail."
        return

    # Verifique se a lista de usuários não é None ou vazia
    if not lista_autenticar:
        self.label_email_autenticar.text = "Erro ao obter a lista de usuários válidos."
        return

    # Certifique-se de que `usuarios_validos` é uma lista e faça a verificação
    if isinstance(lista_autenticar, list) and email in [u.lower() for u in lista_autenticar]:
        self.label_email_autenticar.text = f" Acesso permitido! Bem-vindo(a)!"
        self.botao_entrar.disabled = False  # Desabilita o botão após sucesso na validação
        self.botao_entrar.color = 0, 1 ,0 ,1
    else:
        self.label_email_autenticar.text = "Acesso negado! Usuário não autorizado."
        self.botao_entrar.disabled = True  # Certifica-se de que o botão está habilitado se necessário
        self.botao_entrar.color = 1, 0 ,0 ,1